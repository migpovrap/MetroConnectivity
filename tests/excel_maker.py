import os
import time
import subprocess
import sys
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.chart import ScatterChart, Reference, Series
from openpyxl.chart.trendline import Trendline

def get_test_sizes(test_file):
  with open(test_file, 'r', encoding='utf-8') as file:
    first_line = file.readline().strip()
    num_stations, num_connections, num_lines = map(int, first_line.split()[:3])
    return num_stations, num_connections, num_lines

def run_program(test_file):
  start_time = time.time()
  process = subprocess.Popen(['../src/main'], stdin=open(test_file, 'r', encoding='utf-8'), stdout=subprocess.PIPE, stderr=subprocess.PIPE)
  process.wait()
  end_time = time.time()
  execution_time = end_time - start_time
  return execution_time

def wrap_text(header_text, max_length=30):
  """
  Wrap text to fit in at most two lines, splitting only at spaces.
  Will not break words in the middle.
  """
  words = header_text.split()
  wrapped_lines = []
  current_line = ""

  for word in words:
    # If adding this word exceeds the max_length, start a new line
    if len(current_line) + len(word) + 1 > max_length:
      wrapped_lines.append(current_line)
      current_line = word
    else:
      # Add word to the current line
      if current_line:
        current_line += " " + word
      else:
        current_line = word

  if current_line:
    wrapped_lines.append(current_line)
  return "\n".join(wrapped_lines)

def main():
  tests_folder = os.path.dirname(os.path.abspath(__file__))
  if len(sys.argv) != 4:
    print("Usage: python excel_maker.py <number_of_runs> <num_stations> <num_connections> <num_lines>")
    sys.exit(1)

  num_runs = int(sys.argv[1])
  m_exponent = float(sys.argv[2])
  n_exponent = float(sys.argv[3])

  output_file = "ASA_Relatório.xlsx"

  wb = Workbook()
  ws = wb.active
  ws.title = "Results"

  # Create headers
  headers = ["Test File", "Operation Table Size", "Sequence Size"]
  for run in range(num_runs):
    headers.append(f"Time {run + 1}")
  headers.append("Average Time (s)")
  headers.append("Uncertainty")
  headers.append("Complexity")

  # Wrap headers and write with wrap text
  for col_idx, header in enumerate(headers, start=1):
    wrapped_header = wrap_text(header)  # Wrap the header text
    cell = ws.cell(row=1, column=col_idx)
    cell.value = wrapped_header
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    # Set the font color to white
    cell.font = Font(color="FFFFFF")

  # Set header row height
  ws.row_dimensions[1].height = 40  # Adjust height to fit two lines

  results = []
  for test_file in os.listdir(tests_folder):
    if test_file.endswith('.in'):
      test_path = os.path.join(tests_folder, test_file)
      matrix_size, sequence_size = get_test_sizes(test_path)
      execution_times = [run_program(test_path) for _ in range(num_runs)]
      results.append([test_file, matrix_size, sequence_size] + execution_times)

  # Sort results alphabetically by test file name
  results.sort(key=lambda x: x[0])

  # Write sorted data to the sheet
  for row_idx, result in enumerate(results, start=2):
    for col_idx, value in enumerate(result, start=1):
      cell = ws.cell(row=row_idx, column=col_idx)
      cell.value = value
      cell.alignment = Alignment(horizontal="center", vertical="center")
      if isinstance(value, float):
        cell.number_format = '0.00E+00'  # Format to 3 significant figures

    # Add average time formula
    start_col = len(result) - num_runs + 1  # First column of execution times
    end_col = len(result)  # Last column of execution times
    avg_formula = f"=AVERAGE({chr(64 + start_col)}{row_idx}:{chr(64 + end_col)}{row_idx})"
    avg_cell = ws.cell(row=row_idx, column=len(result) + 1)
    avg_cell.value = avg_formula
    avg_cell.alignment = Alignment(horizontal="center", vertical="center")
    avg_cell.number_format = '0.00E+00'  # Format to 3 significant figures

    # Add standard deviation formula
    stdev_formula = f"=STDEV({chr(64 + start_col)}{row_idx}:{chr(64 + end_col)}{row_idx})"
    stdev_cell = ws.cell(row=row_idx, column=len(result) + 2)
    stdev_cell.value = stdev_formula
    stdev_cell.alignment = Alignment(horizontal="center", vertical="center")
    stdev_cell.number_format = '0.0E+0'  # Format to 1 significant figure

    # Add complexity formula
    complexity_formula = f"=B{row_idx}^{m_exponent}*C{row_idx}^{n_exponent}"
    complexity_cell = ws.cell(row=row_idx, column=len(result) + 3)
    complexity_cell.value = complexity_formula
    complexity_cell.alignment = Alignment(horizontal="center", vertical="center")

  # Define table range
  table_end_col = len(headers)
  table = Table(displayName="ResultsTable", ref=f"A1:{chr(64 + table_end_col)}{len(results) + 1}")

  # Adjust the table style to show row stripes but not column stripes
  style = TableStyleInfo(
    name="TableStyleMedium9",
    showFirstColumn=False,
    showLastColumn=False,
    showRowStripes=True,  # Alternating row colors
    showColumnStripes=False  # No alternating column colors
  )
  table.tableStyleInfo = style
  ws.add_table(table)

  # Set all column widths to 10 and wrap text
  for col in range(1, len(headers) + 1):
    col_letter = chr(64 + col)
    ws.column_dimensions[col_letter].width = 10  # Set column width to 10
    for row in range(1, len(results) + 2):  # Include header row
      cell = ws.cell(row=row, column=col)
      cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

  # Add Scatter Plot
  chart = ScatterChart()
  chart.title = "Complexity vs Average Time"
  chart.x_axis.title = "Complexity"
  chart.y_axis.title = "Average Time (s)"

  # Add major and minor grid lines

  # Show gridlines
  ws.sheet_view.showGridLines = True
  # Tick Marks
  chart.x_axis.majorTickMark = "cross"
  chart.x_axis.minorTickMark = "out"

  # Define data ranges
  result_length = len(results[0]) if results else 0
  complexity_range = Reference(ws, min_col=result_length + 3, min_row=2, max_row=len(results) + 1)
  avg_time_range = Reference(ws, min_col=result_length + 1, min_row=2, max_row=len(results) + 1)

  # Add series to chart
  series = Series(avg_time_range, xvalues=complexity_range)
  chart.series.append(series)

  # Add trendline
  trendline = Trendline(trendlineType="linear")
  series.trendline = trendline

  # Remove the legend
  chart.legend = None

  # Add chart to sheet
  ws.add_chart(chart, f"E{len(results) + 5}")

  # Remove the file if it already exists
  if os.path.exists(output_file):
    os.remove(output_file)

  wb.save(output_file)
  print(f"Results saved to {output_file}")

if __name__ == '__main__':
  main()
